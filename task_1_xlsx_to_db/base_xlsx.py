from typing import Generator

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from settings import DIR_PATH, SKIP_SHEETS
from os import path


class BaseXlsx(object):
    __slots__ = 'file_name', 'file_path', 'xlsx', 'normalize'

    def __init__(self, file_name: str, normalize=True) -> None:
        self.file_name: str = file_name
        self.file_path: str = path.join(DIR_PATH, file_name)
        self.xlsx: Workbook = self.loader(self.file_path)
        self.normalize: bool = normalize
        if self.normalize:
            self.__normalizer()

    @classmethod
    def loader(cls, file_path: str) -> Workbook:
        assert file_path.endswith('.xlsx')
        try:
            work_book: Workbook = load_workbook(file_path)
            return work_book
        except FileNotFoundError:
            raise FileNotFoundError(f'Файл {file_path} не существует')

    def get_sheets(self) -> Generator[Worksheet, None, None]:
        for sheet in self.xlsx.sheetnames:
            if sheet.lower() in SKIP_SHEETS:
                continue
            yield self.xlsx[sheet]

    def __get_sheet_name(self, sheet: str) -> str:
        ...

    def __normalizer(self) -> Worksheet:
        ...

    def get_normalized_sheets(self) -> Generator[Worksheet, None, None]:
        if self.normalize:
            ...
